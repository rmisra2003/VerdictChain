"""Media extraction helpers for evidence intelligence.

DeepSeek's public chat API is text-first, so VerdictChain extracts useful
signals from uploaded media before asking DeepSeek to summarize and structure
the evidence. Every extractor is best-effort: failures become metadata and
warnings instead of breaking chain-of-custody ingestion.
"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
from dataclasses import dataclass, field
from typing import Any

from app.services.openai_media_service import openai_media_service

logger = logging.getLogger(__name__)

MAX_EXTRACTED_CHARS = 30000
MAX_SPREADSHEET_ROWS = 80
MAX_SPREADSHEET_COLS = 12


@dataclass
class ExtractionResult:
    """Structured output from a media extraction pass."""

    media_kind: str
    extraction_status: str
    extracted_text: str = ""
    metadata: dict[str, Any] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    @property
    def text_excerpt(self) -> str:
        return self.extracted_text[:1200]

    def as_metadata(self) -> dict[str, Any]:
        return {
            **self.metadata,
            "warnings": self.warnings,
            "text_length": len(self.extracted_text),
        }


class MediaExtractionService:
    """Best-effort text/OCR/transcription extraction from uploaded files."""

    async def extract(
        self,
        file_data: bytes,
        content_type: str,
        filename: str,
    ) -> ExtractionResult:
        """Extract analyzable text and metadata from an evidence file."""
        normalized_type = (content_type or "application/octet-stream").lower()
        media_kind = self._media_kind(normalized_type, filename)
        metadata = {
            "filename": filename,
            "content_type": normalized_type,
            "size_bytes": len(file_data),
            "extractor": media_kind,
        }

        try:
            if media_kind in {"text", "json", "csv"}:
                text = await asyncio.to_thread(
                    self._extract_text_like,
                    file_data,
                    normalized_type,
                )
                return self._success(media_kind, text, metadata)

            if media_kind == "spreadsheet":
                text = await asyncio.to_thread(self._extract_spreadsheet, file_data)
                return self._success(media_kind, text, metadata)

            if media_kind == "pdf":
                text = await asyncio.to_thread(self._extract_pdf, file_data)
                status = "extracted" if text.strip() else "metadata_only"
                warnings = [] if text.strip() else ["No selectable PDF text found; OCR would be required for scanned pages."]
                return ExtractionResult(media_kind, status, text, metadata, warnings)

            if media_kind == "image":
                text, image_meta, warnings = await self._extract_image(
                    file_data,
                    normalized_type,
                    filename,
                )
                metadata.update(image_meta)
                status = "extracted" if text.strip() else "metadata_only"
                return ExtractionResult(media_kind, status, text, metadata, warnings)

            if media_kind == "audio":
                text, audio_meta, warnings = await openai_media_service.transcribe_audio(
                    file_data,
                    normalized_type,
                    filename,
                )
                metadata.update(audio_meta)
                status = "extracted" if text.strip() else "metadata_only"
                return ExtractionResult(media_kind, status, text, metadata, warnings)

            if media_kind == "video":
                return ExtractionResult(
                    media_kind,
                    "metadata_only",
                    "",
                    metadata,
                    ["Video keyframe extraction is not configured in this deployment."],
                )

            return ExtractionResult(
                media_kind,
                "metadata_only",
                "",
                metadata,
                ["No extractor available for this MIME type."],
            )
        except Exception as exc:
            logger.exception("Media extraction failed for %s", filename)
            return ExtractionResult(
                media_kind,
                "failed",
                "",
                metadata,
                [f"Extraction failed: {exc}"],
            )

    def _success(self, media_kind: str, text: str, metadata: dict[str, Any]) -> ExtractionResult:
        text = self._limit(text)
        return ExtractionResult(
            media_kind=media_kind,
            extraction_status="extracted" if text.strip() else "metadata_only",
            extracted_text=text,
            metadata=metadata,
            warnings=[] if text.strip() else ["No readable text was extracted."],
        )

    def _media_kind(self, content_type: str, filename: str) -> str:
        suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
        if content_type in {"text/plain"} or suffix in {"txt", "md", "log"}:
            return "text"
        if content_type == "application/json" or suffix == "json":
            return "json"
        if content_type == "text/csv" or suffix == "csv":
            return "csv"
        if content_type == "application/pdf" or suffix == "pdf":
            return "pdf"
        if content_type.startswith("image/") or suffix in {"png", "jpg", "jpeg", "webp"}:
            return "image"
        if content_type.startswith("audio/") or suffix in {"wav", "mp3", "mpga", "m4a", "aac", "flac", "ogg"}:
            return "audio"
        if content_type.startswith("video/") or suffix in {"mp4", "mov", "mkv", "webm"}:
            return "video"
        if suffix in {"xlsx", "xlsm"}:
            return "spreadsheet"
        return "binary"

    def _extract_text_like(self, file_data: bytes, content_type: str) -> str:
        raw_text = self._decode(file_data)
        if content_type == "application/json":
            try:
                return json.dumps(json.loads(raw_text), indent=2, sort_keys=True)
            except json.JSONDecodeError:
                return raw_text
        if content_type == "text/csv":
            rows = list(csv.reader(io.StringIO(raw_text)))
            preview = rows[:MAX_SPREADSHEET_ROWS]
            return "\n".join(", ".join(cell.strip() for cell in row) for row in preview)
        return raw_text

    def _extract_pdf(self, file_data: bytes) -> str:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("PDF extractor dependency is not installed.") from exc

        reader = PdfReader(io.BytesIO(file_data))
        parts: list[str] = []
        for index, page in enumerate(reader.pages[:30], start=1):
            text = page.extract_text() or ""
            if text.strip():
                parts.append(f"[Page {index}]\n{text.strip()}")
        return self._limit("\n\n".join(parts))

    def _extract_spreadsheet(self, file_data: bytes) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Spreadsheet extractor dependency is not installed.") from exc

        workbook = load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet in workbook.worksheets[:5]:
            parts.append(f"[Sheet: {sheet.title}]")
            for row in sheet.iter_rows(
                min_row=1,
                max_row=MAX_SPREADSHEET_ROWS,
                max_col=MAX_SPREADSHEET_COLS,
                values_only=True,
            ):
                values = ["" if value is None else str(value) for value in row]
                if any(cell.strip() for cell in values):
                    parts.append(" | ".join(values))
        return self._limit("\n".join(parts))

    async def _extract_image(
        self,
        file_data: bytes,
        content_type: str,
        filename: str,
    ) -> tuple[str, dict[str, Any], list[str]]:
        ocr_text, metadata, warnings = await asyncio.to_thread(self._extract_image_ocr, file_data)
        vision_text, vision_meta, vision_warnings = await openai_media_service.analyze_image(
            file_data,
            content_type,
            filename,
        )
        metadata.update(vision_meta)
        warnings.extend(vision_warnings)

        parts: list[str] = []
        if ocr_text.strip():
            parts.append(f"[Local OCR]\n{ocr_text.strip()}")
        if vision_text.strip():
            parts.append(f"[OpenAI Vision]\n{vision_text.strip()}")

        combined_text = self._limit("\n\n".join(parts))
        if not combined_text and not warnings:
            warnings.append("No readable image text or visual summary was extracted.")
        return combined_text, metadata, warnings

    def _extract_image_ocr(self, file_data: bytes) -> tuple[str, dict[str, Any], list[str]]:
        warnings: list[str] = []
        metadata: dict[str, Any] = {}
        try:
            from PIL import Image
        except ImportError as exc:
            raise RuntimeError("Image extractor dependency is not installed.") from exc

        image = Image.open(io.BytesIO(file_data))
        metadata.update({
            "image_format": image.format,
            "image_width": image.width,
            "image_height": image.height,
            "image_mode": image.mode,
        })

        try:
            import pytesseract

            prepared_image = self._prepare_image_for_ocr(image)
            text = pytesseract.image_to_string(prepared_image)
            metadata["ocr_engine"] = "tesseract"
            return self._limit(text), metadata, warnings
        except Exception as exc:
            warnings.append(
                "Image OCR is unavailable in this runtime; install the Tesseract binary to enable it."
            )
            warnings.append(str(exc))
            return "", metadata, warnings

    def _prepare_image_for_ocr(self, image) -> Any:
        """Increase OCR contrast and scale small images before Tesseract."""
        from PIL import ImageOps

        prepared = image.convert("L")
        width, height = prepared.size
        scale = 2 if max(width, height) < 1600 else 1
        if scale > 1:
            prepared = prepared.resize((width * scale, height * scale))
        prepared = ImageOps.autocontrast(prepared)
        return prepared

    def _decode(self, file_data: bytes) -> str:
        for encoding in ("utf-8", "utf-16", "latin-1"):
            try:
                return file_data.decode(encoding)
            except UnicodeDecodeError:
                continue
        return file_data.decode("utf-8", errors="ignore")

    def _limit(self, text: str) -> str:
        return text.strip()[:MAX_EXTRACTED_CHARS]


media_extraction_service = MediaExtractionService()
